from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import olefile
from openpyxl import load_workbook
from pypdf import PdfReader
from PIL import Image, UnidentifiedImageError
from sqlalchemy.orm import Session

from ..enums import AttachmentStatus
from ..models import Attachment

TEXT_EXTENSIONS = {".txt", ".csv", ".log", ".md"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}


def _sniff_image_format(path: Path) -> str | None:
    """확장자가 없거나 신뢰할 수 없을 때, 실제 파일 내용을 열어 이미지인지 확인한다.

    다음 IMAP에서 내려받은 '전달 이미지' 같은 첨부는 파일명에 확장자가 없고
    content_type도 실제와 다른 경우가 있어(예: 실제 JPEG인데 image/png로 기록),
    확장자 대신 파일 내용을 직접 열어봐서 판단한다.
    """
    try:
        with Image.open(path) as image:
            return (image.format or "").lower() or None
    except (UnidentifiedImageError, OSError):
        return None


def _read_text_file(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_hwpx_text(path: Path) -> str:
    text_parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        names = sorted(
            name
            for name in archive.namelist()
            if name.lower().endswith(".xml")
            and ("contents/section" in name.lower() or "header.xml" in name.lower())
        )
        for name in names:
            root = ElementTree.fromstring(archive.read(name))
            for element in root.iter():
                tag = element.tag.rsplit("}", 1)[-1].lower()
                if tag in {"t", "text"} and element.text:
                    text_parts.append(element.text)
    return "\n".join(part.strip() for part in text_parts if part.strip())


def extract_pptx_text(path: Path) -> str:
    slides: list[tuple[int, str]] = []
    with zipfile.ZipFile(path) as archive:
        slide_names = [
            (int(match.group(1)), name)
            for name in archive.namelist()
            if (match := re.match(r"ppt/slides/slide(\d+)\.xml$", name))
        ]
        for index, name in sorted(slide_names):
            root = ElementTree.fromstring(archive.read(name))
            runs = [
                (element.text or "").strip()
                for element in root.iter()
                if element.tag.rsplit("}", 1)[-1] == "t" and (element.text or "").strip()
            ]
            if runs:
                slides.append((index, " ".join(runs)))
    return "\n\n".join(f"[슬라이드 {index}]\n{text}" for index, text in slides)


def extract_hwp_preview_text(path: Path) -> str:
    """HWP(5.0 바이너리) 문서의 PrvText(미리보기) 스트림에서 텍스트를 뽑는다.

    한글이 검색·미리보기용으로 저장해두는 요약 텍스트를 읽는 것이라 본문 전체를
    파싱하는 것보다 정확도가 낮다. 표/도형 안의 텍스트나 긴 문서의 뒷부분은
    빠질 수 있으니 호출 측에서 참고용으로만 다뤄야 한다.
    """
    if not olefile.isOleFile(str(path)):
        return ""
    with olefile.OleFileIO(str(path)) as ole:
        if not ole.exists("PrvText"):
            return ""
        raw = ole.openstream("PrvText").read()
    return raw.decode("utf-16-le", errors="replace").strip()


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    return "\n\n".join(page for page in pages if page)


def extract_excel_text(path: Path, max_cells: int = 1500) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    output: list[str] = []
    cell_count = 0
    try:
        for sheet in workbook.worksheets:
            output.append(f"[시트: {sheet.title}]")
            for row in sheet.iter_rows():
                values = [str(cell.value).strip() for cell in row if cell.value not in (None, "")]
                if values:
                    output.append(" | ".join(values))
                    cell_count += len(values)
                if cell_count >= max_cells:
                    output.append("[이하 생략]")
                    return "\n".join(output)
    finally:
        workbook.close()
    return "\n".join(output)


def process_attachment(session: Session, attachment: Attachment) -> Attachment:
    path = Path(attachment.saved_path)
    extension = path.suffix.lower()
    try:
        if extension == ".hwpx":
            attachment.extracted_text = extract_hwpx_text(path)
            attachment.status = AttachmentStatus.EXTRACTED
        elif extension == ".hwp":
            preview_text = extract_hwp_preview_text(path)
            if preview_text.strip():
                attachment.extracted_text = preview_text
                attachment.status = AttachmentStatus.EXTRACTED
                attachment.error_message = (
                    "미리보기(PrvText) 기반 추출이라 표/도형 안의 텍스트나 "
                    "문서 뒷부분 내용은 빠졌을 수 있습니다."
                )
            else:
                attachment.status = AttachmentStatus.MANUAL_REVIEW
                attachment.error_message = (
                    "HWP 미리보기 텍스트를 찾지 못했습니다. "
                    "HWPX/PDF로 변환하거나 한컴오피스 자동화 연동이 필요합니다."
                )
        elif extension == ".pptx":
            attachment.extracted_text = extract_pptx_text(path)
            if attachment.extracted_text.strip():
                attachment.status = AttachmentStatus.EXTRACTED
            else:
                attachment.status = AttachmentStatus.MANUAL_REVIEW
                attachment.error_message = "슬라이드에서 텍스트를 찾지 못했습니다."
        elif extension == ".pdf":
            if not path.read_bytes()[:8].startswith(b"%PDF-"):
                # 확장자는 .pdf지만 실제 PDF 헤더가 아님 — 문서보안(DRM) 등으로
                # 암호화된 파일일 가능성이 높다. pypdf에 던져봐야 알아보기 힘든
                # 파싱 에러만 남으므로 미리 걸러서 원인을 명확히 남긴다.
                attachment.status = AttachmentStatus.FAILED
                attachment.error_message = (
                    "PDF 형식이 아닙니다. 문서보안(DRM) 솔루션으로 암호화된 파일일 수 있어 "
                    "발신자에게 보안이 걸리지 않은 원본을 다시 요청해야 할 수 있습니다."
                )
            else:
                attachment.extracted_text = extract_pdf_text(path)
                if attachment.extracted_text.strip():
                    attachment.status = AttachmentStatus.EXTRACTED
                else:
                    attachment.status = AttachmentStatus.IMAGE_PENDING
                    attachment.error_message = "스캔 PDF로 판단되어 이미지 분석이 필요합니다."
        elif extension in {".xlsx", ".xlsm"}:
            attachment.extracted_text = extract_excel_text(path)
            attachment.status = AttachmentStatus.EXTRACTED
        elif extension in IMAGE_EXTENSIONS:
            attachment.status = AttachmentStatus.IMAGE_PENDING
        elif extension in TEXT_EXTENSIONS:
            attachment.extracted_text = _read_text_file(path)
            attachment.status = AttachmentStatus.EXTRACTED
        elif _sniff_image_format(path) is not None:
            # 확장자가 없거나 신뢰할 수 없는 첨부(예: 전달 이미지)를 내용 기반으로 구제.
            attachment.status = AttachmentStatus.IMAGE_PENDING
            attachment.error_message = "확장자가 없어 파일 내용을 분석해 이미지로 판별했습니다."
        else:
            attachment.status = AttachmentStatus.MANUAL_REVIEW
            attachment.error_message = f"지원하지 않는 첨부 형식: {extension or '확장자 없음'}"
    except Exception as error:  # 파일별 실패를 전체 메일 실패로 전파하지 않는다.
        attachment.status = AttachmentStatus.FAILED
        attachment.error_message = f"{type(error).__name__}: {error}"
    session.add(attachment)
    session.flush()
    return attachment


def compact_attachment_context(attachments: list[Attachment], max_chars: int = 12_000) -> str:
    blocks: list[str] = []
    remaining = max_chars
    for attachment in attachments:
        text = (attachment.extracted_text or attachment.analysis_summary or "").strip()
        if not text:
            continue
        block = f"\n[첨부파일: {attachment.filename}]\n{text}"
        if len(block) > remaining:
            block = block[:remaining]
        blocks.append(block)
        remaining -= len(block)
        if remaining <= 0:
            break
    return "\n".join(blocks)
