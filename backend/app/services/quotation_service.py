from __future__ import annotations

import copy
import json
import os
import re
import shutil
import subprocess
import sys
import uuid
import zipfile
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from ..config import PROJECT_ROOT, Settings
from ..enums import DraftStatus, MailStatus, Severity
from ..models import Mail, QuotationDraft, QuotationDraftItem, ReviewIssue
from .price_engine_adapter import calculate_item_price
from .quote_math import validate_quote_items
from .utils import sanitize_filename

StorageMode = Literal["existing", "department", "person", "separate"]
MAIL_MARKER_PREFIX = "OPENMOON_MAIL_ID:"
ITEM_START_ROW = 14
ITEM_END_ROW = 23

EXCEL_COM_SCRIPT = r'''
param([string]$TargetPath, [string]$TemplatePath, [string]$TemplateSheet, [string]$PayloadPath)
$ErrorActionPreference = "Stop"
function Set-ExcelValue($Range, $Value) {
    if ($null -eq $Value) { $Range.ClearContents(); return }
    if ($Value -is [byte] -or $Value -is [int16] -or $Value -is [int32] -or $Value -is [int64] -or $Value -is [single] -or $Value -is [double] -or $Value -is [decimal]) {
        $number = [Convert]::ToString($Value, [Globalization.CultureInfo]::InvariantCulture)
        $Range.Formula = "=" + $number
        return
    }
    $Range.Value2 = $Value.ToString()
}
$excel = $null
$targetBook = $null
$templateBook = $null
try {
    $payload = Get-Content -LiteralPath $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $targetBook = $excel.Workbooks.Open($TargetPath, 0, $false)
    $templateBook = $excel.Workbooks.Open($TemplatePath, 0, $true)

    $oldSheet = $null
    foreach ($candidate in @($targetBook.Worksheets)) {
        $found = $candidate.Cells.Find($payload.marker)
        if ($null -ne $found) { $oldSheet = $candidate; break }
    }
    $oldName = $null
    $oldIndex = $null
    if ($null -ne $oldSheet) {
        $oldName = $oldSheet.Name
        $oldIndex = $oldSheet.Index
        $oldSheet.Delete()
    }

    $sourceSheet = $templateBook.Worksheets.Item($TemplateSheet)
    $sourceSheet.Copy([Type]::Missing, $targetBook.Worksheets.Item($targetBook.Worksheets.Count))
    $sheet = $targetBook.ActiveSheet
    if ($null -ne $oldIndex -and $oldIndex -le $targetBook.Worksheets.Count) {
        $sheet.Move($targetBook.Worksheets.Item($oldIndex))
        $sheet = $targetBook.Worksheets.Item($oldIndex)
    }

    $baseName = if ($null -ne $oldName) { $oldName } else { [string]$payload.base_name }
    $name = $baseName
    if ($null -eq $oldName) {
        $number = 2
        while ($true) {
            $exists = $false
            foreach ($existingSheet in @($targetBook.Worksheets)) {
                if ($existingSheet.Name -eq $name -and $existingSheet.Index -ne $sheet.Index) { $exists = $true; break }
            }
            if (-not $exists) { break }
            $suffix = "_" + $number
            $prefixLength = [Math]::Min(31 - $suffix.Length, $baseName.Length)
            $name = $baseName.Substring(0, $prefixLength) + $suffix
            $number++
        }
    }
    $sheet.Name = $name

    foreach ($entry in $payload.cells.PSObject.Properties) {
        $range = $sheet.Range($entry.Name)
        if ($range.MergeCells) { $range = $range.MergeArea.Cells.Item(1, 1) }
        if ($entry.Name -eq "D5") {
            $dateValue = if ($entry.Value -is [datetime]) { $entry.Value } else { [datetime]::Parse($entry.Value.ToString()) }
            # Some Excel/PowerShell COM combinations expose Value/Value2 as a
            # string-only setter. A formula string avoids that interop cast
            # while retaining the template cell's date number format.
            $range.Formula = "=DATE($($dateValue.Year),$($dateValue.Month),$($dateValue.Day))"
        }
        else { Set-ExcelValue $range $entry.Value }
    }

    for ($row = 14; $row -le 23; $row++) {
        foreach ($column in @(2, 3, 6, 7, 9, 12, 20)) {
            $range = $sheet.Cells.Item($row, $column)
            if ($range.MergeCells) {
                # Excel refuses ClearContents on only one cell of a merged
                # range. These configured columns are the top-left anchors,
                # so clear the complete merge area exactly once.
                $range.MergeArea.ClearContents()
            }
            else { $range.ClearContents() }
        }
    }

    foreach ($item in @($payload.items)) {
        $row = [int]$item.row
        Set-ExcelValue ($sheet.Cells.Item($row, 2)) ([int]$item.position)
        $text = [string]$item.product
        if (-not [string]::IsNullOrWhiteSpace([string]$item.detail)) { $text += "`n(" + [string]$item.detail + ")" }
        $itemRange = $sheet.Cells.Item($row, 3)
        if ($itemRange.MergeCells) { $itemRange = $itemRange.MergeArea.Cells.Item(1, 1) }
        $itemRange.Value2 = $text
        $itemRange.Characters(1, ([string]$item.product).Length).Font.Bold = $true
        $itemRange.Characters(1, ([string]$item.product).Length).Font.Size = 14
        if ($text.Length -gt ([string]$item.product).Length) {
            $start = ([string]$item.product).Length + 2
            $length = $text.Length - $start + 1
            $itemRange.Characters($start, $length).Font.Bold = $false
            $itemRange.Characters($start, $length).Font.Size = 9
        }
        Set-ExcelValue ($sheet.Cells.Item($row, 6)) $item.quantity
        Set-ExcelValue ($sheet.Cells.Item($row, 7)) $item.unit_price
        Set-ExcelValue ($sheet.Cells.Item($row, 9)) $item.amount
        Set-ExcelValue ($sheet.Cells.Item($row, 12)) $item.note
        Set-ExcelValue ($sheet.Cells.Item($row, 20)) $item.cost_price
    }

    # T열은 제작원가 및 내부 식별정보 전용이며 고객 화면에서는 숨긴다.
    $sheet.Columns.Item(20).Hidden = $true

    $markerCell = $sheet.Cells.Item(1, 20)
    Set-ExcelValue $markerCell $payload.marker
    $markerCell.EntireColumn.Hidden = $true
    $targetBook.Save()
}
finally {
    if ($null -ne $templateBook) { $templateBook.Close($false) }
    if ($null -ne $targetBook) { $targetBook.Close($false) }
    if ($null -ne $excel) { $excel.Quit() }
    if ($null -ne $sourceSheet) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($sourceSheet) }
    if ($null -ne $sheet) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($sheet) }
    if ($null -ne $templateBook) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($templateBook) }
    if ($null -ne $targetBook) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($targetBook) }
    if ($null -ne $excel) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
'''


CUSTOMER_PRIVATE_CELLS = ("L5", "L6", "L7")


def _catalog_product_names() -> list[str]:
    path = (
        Path(__file__).resolve().parents[3]
        / "config"
        / "product_catalog.json"
    )

    try:
        data = json.loads(
            path.read_text(encoding="utf-8")
        )
    except (
        OSError,
        ValueError,
        TypeError,
    ):
        return []

    rows: list[Any]

    if isinstance(data, dict):
        candidate = data.get("products", [])
        rows = (
            candidate
            if isinstance(candidate, list)
            else []
        )
    elif isinstance(data, list):
        rows = data
    else:
        rows = []

    names: list[str] = []

    for row in rows:
        if not isinstance(row, dict):
            continue

        name = str(
            row.get("name")
            or ""
        ).strip()

        if name:
            names.append(name)

    return names


def _canonical_customer_product_name(
    raw_name: str | None,
    normalized_product: str | None = None,
    catalog_names: list[str] | None = None,
) -> str:
    raw = (
        raw_name
        or ""
    ).strip()

    normalized = (
        normalized_product
        or ""
    ).strip()

    # 분석/카탈로그 매칭 단계에서 이미 대표품목이 확정됐다면 우선 사용한다.
    if (
        normalized
        and normalized != raw
    ):
        return normalized

    if not raw:
        return "품목"

    names = (
        catalog_names
        if catalog_names is not None
        else _catalog_product_names()
    )

    def compact(value: str) -> str:
        return re.sub(
            r"[^0-9A-Za-z가-힣]",
            "",
            value,
        ).casefold()

    raw_key = compact(raw)

    # 정확히 카탈로그 품목명이면 그대로 유지.
    for name in names:
        if compact(name) == raw_key:
            return name

    # "친환경 현수막", "실내용 현수막"처럼 수식어가 붙은 경우
    # 카탈로그의 대표 품목명이 원문 안에 있으면 대표명으로 축약한다.
    contained = [
        name
        for name in names
        if compact(name)
        and compact(name) in raw_key
    ]

    if contained:
        # 가장 구체적인 대표품목명을 우선.
        return max(
            contained,
            key=lambda value: len(
                compact(value)
            ),
        )

    # 회사 요구사항의 대표 예시를 최종 안전망으로 처리.
    if "현수막" in raw:
        return "현수막"

    return raw


CUSTOMER_PDF_EXPORT_SCRIPT = r"""
param(
    [string]$SourcePath,
    [string]$PdfPath,
    [string]$Marker,
    [string]$PayloadPath
)

$ErrorActionPreference = "Stop"
$excel = $null
$workbook = $null
$sheet = $null

try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false

    $workbook = $excel.Workbooks.Open(
        $SourcePath,
        0,
        $false
    )

    foreach ($candidate in @($workbook.Worksheets)) {
        $found = $candidate.Cells.Find($Marker)
        if ($null -ne $found) {
            $sheet = $candidate
            break
        }
    }

    if ($null -eq $sheet) {
        throw "고객용 PDF로 내보낼 견적 시트를 찾지 못했습니다."
    }

    foreach ($address in @("L5", "L6", "L7")) {
        $range = $sheet.Range($address)
        if ($range.MergeCells) {
            $range.MergeArea.ClearContents()
        }
        else {
            $range.ClearContents()
        }
    }

    $sheet.Columns.Item(20).ClearContents()
    $sheet.Columns.Item(20).Hidden = $true

    # 고객용 PDF의 품목명은 내부 분석명이 아니라 대표 품목명으로 표시한다.
    if (
        -not [string]::IsNullOrWhiteSpace($PayloadPath)
        -and (Test-Path -LiteralPath $PayloadPath)
    ) {
        $payload = (
            Get-Content
                -LiteralPath $PayloadPath
                -Raw
                -Encoding UTF8
            | ConvertFrom-Json
        )

        foreach ($item in @($payload.items)) {
            $row = [int]$item.row
            $product = [string]$item.product

            if ([string]::IsNullOrWhiteSpace($product)) {
                continue
            }

            $range = $sheet.Cells.Item(
                $row,
                3
            )

            if ($range.MergeCells) {
                $range = $range.MergeArea.Cells.Item(
                    1,
                    1
                )
            }

            $current = [string]$range.Value2
            $detail = ""

            if (
                -not [string]::IsNullOrWhiteSpace($current)
                -and $current.Contains("`n")
            ) {
                $parts = $current -split "`n", 2

                if ($parts.Count -gt 1) {
                    $detail = [string]$parts[1]
                }
            }

            $newText = $product

            if (
                -not [string]::IsNullOrWhiteSpace($detail)
            ) {
                $newText += "`n" + $detail
            }

            $range.Value2 = $newText

            try {
                $range.Characters(
                    1,
                    $product.Length
                ).Font.Bold = $true

                $range.Characters(
                    1,
                    $product.Length
                ).Font.Size = 14

                if ($newText.Length -gt $product.Length) {
                    $start = $product.Length + 2
                    $length = $newText.Length - $start + 1

                    $range.Characters(
                        $start,
                        $length
                    ).Font.Bold = $false

                    $range.Characters(
                        $start,
                        $length
                    ).Font.Size = 9
                }
            }
            catch {
                # PDF 내용 자체가 우선이므로 부분 서식 실패는 무시한다.
            }
        }
    }

    $sheet.ExportAsFixedFormat(
        0,
        $PdfPath,
        0,
        $true,
        $false
    )
}
finally {
    if ($null -ne $workbook) {
        $workbook.Close($false)
    }
    if ($null -ne $excel) {
        $excel.Quit()
    }

    if ($null -ne $sheet) {
        [void][Runtime.InteropServices.Marshal]::ReleaseComObject($sheet)
    }
    if ($null -ne $workbook) {
        [void][Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)
    }
    if ($null -ne $excel) {
        [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
    }

    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""


def customer_pdf_path(
    settings: Settings,
    draft: QuotationDraft,
) -> Path:
    root = settings.generated_quotes_dir.resolve()
    root.mkdir(parents=True, exist_ok=True)

    customer = _safe_name(
        draft.customer_name,
        "고객",
    )

    filename = sanitize_filename(
        f"견적서_{customer}_{draft.id}.pdf",
        180,
    )

    return (root / filename).resolve()


def _export_customer_pdf(
    settings: Settings,
    draft: QuotationDraft,
    source_path: Path,
    mail: Mail,
) -> Path:
    if sys.platform != "win32":
        raise RuntimeError(
            "고객용 PDF 생성은 현재 Windows용 Microsoft Excel 자동화를 사용합니다."
        )

    if not source_path.exists():
        raise FileNotFoundError(
            f"PDF 원본 견적 XLSX가 없습니다: {source_path}"
        )

    pdf_path = customer_pdf_path(
        settings,
        draft,
    )

    token = uuid.uuid4().hex
    customer_xlsx = pdf_path.with_name(
        f".{pdf_path.stem}.{token}.customer.xlsx"
    )
    temp_pdf = pdf_path.with_name(
        f".{pdf_path.stem}.{token}.saving.pdf"
    )
    script_path = pdf_path.with_name(
        f".{pdf_path.stem}.{token}.pdf-export.ps1"
    )
    payload_path = pdf_path.with_name(
        f".{pdf_path.stem}.{token}.customer.json"
    )

    try:
        shutil.copy2(
            source_path,
            customer_xlsx,
        )

        script_path.write_text(
            CUSTOMER_PDF_EXPORT_SCRIPT,
            encoding="utf-8-sig",
        )

        catalog_names = _catalog_product_names()

        payload = {
            "items": [
                {
                    "row": ITEM_START_ROW + index,
                    "product": _canonical_customer_product_name(
                        getattr(item, "product_name", None),
                        getattr(item, "normalized_product", None),
                        catalog_names,
                    ),
                }
                for index, item in enumerate(mail.items)
                if ITEM_START_ROW + index <= ITEM_END_ROW
            ]
        }

        payload_path.write_text(
            json.dumps(
                payload,
                ensure_ascii=False,
            ),
            encoding="utf-8-sig",
        )

        result = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-NonInteractive",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-SourcePath",
                str(customer_xlsx),
                "-PdfPath",
                str(temp_pdf),
                "-Marker",
                f"{MAIL_MARKER_PREFIX}{mail.id}",
                "-PayloadPath",
                str(payload_path),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
            check=False,
        )

        if result.returncode != 0:
            detail = (
                result.stderr
                or result.stdout
                or "Microsoft Excel PDF 변환에 실패했습니다."
            ).strip()

            raise RuntimeError(
                "고객용 PDF를 생성하지 못했습니다. "
                "Microsoft Excel 설치 및 실행 상태를 확인해주세요.\n"
                + detail
            )

        if not temp_pdf.exists():
            raise RuntimeError(
                "Excel 변환은 종료됐지만 고객용 PDF 파일이 생성되지 않았습니다."
            )

        try:
            os.replace(
                temp_pdf,
                pdf_path,
            )
        except PermissionError as error:
            raise QuotationFileLockedError(
                "고객용 PDF 파일이 다른 프로그램에서 열려 있습니다.\n"
                "PDF를 닫은 뒤 다시 견적서를 생성해주세요."
            ) from error

        return pdf_path

    except FileNotFoundError as error:
        raise RuntimeError(
            "고객용 PDF를 만들려면 Microsoft Excel과 PowerShell이 필요합니다."
        ) from error

    finally:
        customer_xlsx.unlink(missing_ok=True)
        temp_pdf.unlink(missing_ok=True)
        script_path.unlink(missing_ok=True)
        payload_path.unlink(missing_ok=True)


class QuotationFileLockedError(PermissionError):
    pass


def _safe_name(value: str | None, fallback: str) -> str:
    cleaned = sanitize_filename((value or "").strip(), 80).strip(" .-_")
    return cleaned or fallback


def _normalized(value: str | None) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", value or "").casefold()


@lru_cache(maxsize=1)
def _catalog_spec_map() -> dict[str, dict[str, tuple[str, str | None]]]:
    path = PROJECT_ROOT / "config" / "product_catalog.json"

    try:
        with path.open("r", encoding="utf-8") as handle:
            catalog = json.load(handle)
    except (OSError, ValueError):
        return {}

    result: dict[str, dict[str, tuple[str, str | None]]] = {}

    for category in catalog.get("categories", []):
        for product in category.get("products", []):
            fields = {
                str(field.get("key")): (
                    str(field.get("label") or field.get("key")),
                    (
                        str(field.get("legacy_field"))
                        if field.get("legacy_field")
                        else None
                    ),
                )
                for field in product.get("fields", [])
                if field.get("key")
            }

            names = [
                product.get("name"),
                *(product.get("aliases") or []),
            ]

            for name in names:
                normalized = _normalized(str(name or ""))
                if normalized:
                    result[normalized] = fields

    return result


def _dynamic_spec_details(item: Any) -> list[str]:
    attributes = getattr(item, "spec_attributes", None) or {}

    if not isinstance(attributes, dict) or not attributes:
        return []

    product_key = _normalized(
        str(
            getattr(item, "normalized_product", None)
            or getattr(item, "product_name", "")
            or ""
        )
    )

    field_map = _catalog_spec_map().get(product_key, {})
    details: list[str] = []

    duplicate_legacy_fields = {
        "specification",
        "quantity",
        "paper",
        "print_sides",
        "material",
    }

    for key, raw_value in attributes.items():
        if raw_value in (None, "", []):
            continue

        label, legacy_field = field_map.get(
            str(key),
            (str(key), None),
        )

        if legacy_field in duplicate_legacy_fields:
            continue

        if isinstance(raw_value, list):
            value = ", ".join(
                str(part).strip()
                for part in raw_value
                if str(part).strip()
            )
        else:
            value = str(raw_value).strip()

        if value:
            details.append(f"{label}: {value}")

    return details


def _customer_parts(mail: Mail) -> tuple[str, str, str]:
    return (
        _safe_name(mail.customer_organization or mail.customer_name, "고객"),
        _safe_name(mail.customer_department, ""),
        _safe_name(mail.customer_name, "담당자"),
    )


def _new_file_candidates(settings: Settings, mail: Mail, now: datetime) -> list[dict[str, Any]]:
    root = settings.quotation_files_path.resolve()
    company, department, person = _customer_parts(mail)
    year = now.strftime("%y")
    date = now.strftime("%Y%m%d")
    department_stem = f"{year}-{company}{f' {department}' if department else ''}"
    person_stem = f"{department_stem}-{person}" if person else department_stem
    separate_stem = f"견적서_{company}_{person}_{date}"
    rows = [
        ("department", f"{department_stem}.xlsx", "부서 공용 파일"),
        ("person", f"{person_stem}.xlsx", "담당자별 파일"),
        ("separate", f"{separate_stem}.xlsx", "별도 견적 파일"),
    ]
    return [
        {
            "mode": mode,
            "filename": sanitize_filename(filename, 180),
            "file_type": label,
            "exists": (root / sanitize_filename(filename, 180)).exists(),
            "path": str((root / sanitize_filename(filename, 180)).resolve()),
            "related": True,
        }
        for mode, filename, label in rows
    ]


def get_storage_options(
    settings: Settings,
    mail: Mail,
    *,
    now: datetime | None = None,
) -> dict[str, Any]:
    now = now or datetime.now()
    root = settings.quotation_files_path.resolve()
    root.mkdir(parents=True, exist_ok=True)
    company, department, person = _customer_parts(mail)
    company_key = _normalized(company)
    department_key = _normalized(department)
    person_key = _normalized(person)
    existing: list[tuple[int, dict[str, Any]]] = []

    for path in root.glob("*.xlsx"):
        if path.name.startswith((".", "~$")):
            continue
        stem_key = _normalized(path.stem)
        if company_key and company_key not in stem_key:
            continue
        score = 10
        if department_key and department_key in stem_key:
            score += 5
        if person_key and person_key in stem_key:
            score += 8
            file_type = "담당자별 파일"
        elif department_key and department_key in stem_key:
            file_type = "부서 공용 파일"
        else:
            file_type = "기존 관련 파일"
        existing.append(
            (
                score,
                {
                    "mode": "existing",
                    "filename": path.name,
                    "file_type": file_type,
                    "exists": True,
                    "path": str(path.resolve()),
                    "related": True,
                },
            )
        )

    existing.sort(key=lambda row: (-row[0], row[1]["filename"]))
    draft = getattr(mail, "drafts", None)
    selected = str(Path(draft[-1].file_path).resolve()) if draft else None
    return {
        "root_path": str(root),
        "selected_file": selected,
        "existing_files": [row for _, row in existing],
        "new_files": _new_file_candidates(settings, mail, now),
    }


def _validate_target(
    settings: Settings,
    mail: Mail,
    storage_mode: StorageMode,
    target_path: Path,
    now: datetime,
) -> tuple[Path, bool]:
    root = settings.quotation_files_path.resolve()
    target = target_path.resolve()
    try:
        target.relative_to(root)
    except ValueError as error:
        raise ValueError("quotation_files 폴더 밖의 파일은 선택할 수 없습니다.") from error
    if target.suffix.lower() != ".xlsx":
        raise ValueError("XLSX 견적 파일만 선택할 수 있습니다.")
    if storage_mode == "existing":
        if not target.is_file():
            raise FileNotFoundError("선택한 기존 견적 파일을 찾을 수 없습니다.")
        return target, True
    allowed = {
        row["mode"]: Path(row["path"]).resolve()
        for row in _new_file_candidates(settings, mail, now)
    }
    if allowed.get(storage_mode) != target:
        raise ValueError("선택한 저장 방식과 파일 경로가 일치하지 않습니다.")
    return target, target.exists()


def _ensure_no_blocking_reviews(session: Session, mail: Mail) -> None:
    blocking = session.scalar(
        select(ReviewIssue.id).where(
            ReviewIssue.mail_id == mail.id,
            ReviewIssue.resolved.is_(False),
            ReviewIssue.severity == Severity.BLOCKING,
        )
    )
    if blocking:
        raise ValueError("검토가 필요한 필수 항목이 남아 있어 견적서를 생성할 수 없습니다.")


def _resolve_item_price(settings: Settings, mail: Mail, item: Any) -> tuple[int | None, int | None, dict[str, Any]]:
    if item.unit_price is not None:
        unit_price = int(item.unit_price)
        amount = int(item.amount) if item.amount is not None else None
        if amount is None and item.quantity is not None:
            amount = int(round(float(item.quantity) * unit_price))
        source: dict[str, Any] = {
            "type": "MANUAL" if item.confirmed else "MAIL",
            "reason": "담당자가 확정한 단가" if item.confirmed else "메일 또는 분석 결과의 단가",
        }
        evidence = dict(item.evidence or {}).get("price")
        if isinstance(evidence, dict):
            source.update(evidence)
        return unit_price, amount, source
    decision = calculate_item_price(settings=settings, mail=mail, item=item)
    if decision.unit_price is None:
        return None, None, {
            "type": "UNRESOLVED",
            "source": decision.source,
            "reference": decision.reference,
            "score": decision.score,
            "reason": decision.reason,
            "needs_review": True,
        }
    unit_price = int(decision.unit_price)
    amount = int(decision.amount) if decision.amount is not None else None
    if amount is None and item.quantity is not None:
        amount = int(round(float(item.quantity) * unit_price))
    return unit_price, amount, {
        "type": decision.source.upper(),
        "source": decision.source,
        "reference": decision.reference,
        "score": decision.score,
        "reason": decision.reason,
        "needs_review": decision.needs_review,
    }


def _copy_template_sheet(source, target_workbook, title: str, index: int | None = None):
    """Copy a sheet across workbooks without copying private style IDs."""
    target = target_workbook.create_sheet(title=title, index=index)
    for row in source.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target.cell(source_cell.row, source_cell.column)
            target_cell.value = copy.copy(source_cell.value)
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.alignment = copy.copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy.copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy.copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy.copy(source_cell.comment)
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height
        target.row_dimensions[key].hidden = dimension.hidden
        target.row_dimensions[key].outlineLevel = dimension.outlineLevel
        target.row_dimensions[key].collapsed = dimension.collapsed
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
        target.column_dimensions[key].bestFit = dimension.bestFit
        target.column_dimensions[key].outlineLevel = dimension.outlineLevel
        target.column_dimensions[key].collapsed = dimension.collapsed
    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))
    for image in source._images:
        cloned = copy.copy(image)
        cloned.anchor = copy.deepcopy(image.anchor)
        target.add_image(cloned)
    target.sheet_format = copy.copy(source.sheet_format)
    target.sheet_properties = copy.copy(source.sheet_properties)
    target.page_margins = copy.copy(source.page_margins)
    target.page_setup = copy.copy(source.page_setup)
    target.print_options = copy.copy(source.print_options)
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.sheet_view.zoomScale = source.sheet_view.zoomScale
    target.sheet_view.zoomScaleNormal = source.sheet_view.zoomScaleNormal
    target.sheet_view.view = source.sheet_view.view
    target.freeze_panes = source.freeze_panes
    target.auto_filter.ref = source.auto_filter.ref
    target.sheet_state = "visible"
    if source.print_area:
        target.print_area = str(source.print_area).split("!", 1)[-1]
    if source.print_title_rows:
        target.print_title_rows = source.print_title_rows
    if source.print_title_cols:
        target.print_title_cols = source.print_title_cols
    return target


def _merged_anchor(sheet, row: int, column: int):
    cell = sheet.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            return sheet.cell(merged.min_row, merged.min_col)
    raise ValueError(f"병합 셀 시작 위치를 찾을 수 없습니다: {cell.coordinate}")


def _set_value(sheet, coordinate: str, value: Any) -> None:
    cell = sheet[coordinate]
    if isinstance(cell, MergedCell):
        cell = _merged_anchor(sheet, cell.row, cell.column)
    cell.value = value


def _clear_item_area(sheet) -> None:
    for row in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        for column in (2, 3, 6, 7, 9, 12, 20):
            cell = sheet.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _detail_text(item: Any) -> str:
    details: list[str] = []
    specification = str(item.specification or "").strip()
    if specification:
        details.append(specification)
    width_mm = getattr(item, "width_mm", None)
    height_mm = getattr(item, "height_mm", None)
    width_text = f"{width_mm:g}" if isinstance(width_mm, float) else str(width_mm or "")
    height_text = f"{height_mm:g}" if isinstance(height_mm, float) else str(height_mm or "")
    if width_mm is not None and height_mm is not None:
        details.append(f"{width_text}*{height_text}mm")
    elif width_mm is not None:
        details.append(f"가로 {width_text}mm")
    elif height_mm is not None:
        details.append(f"세로 {height_text}mm")
    if item.quantity is not None:
        quantity_text = f"{item.quantity:g}" if isinstance(item.quantity, float) else str(item.quantity)
        details.append(f"{quantity_text}{item.unit or ''}")
    paper = getattr(item, "paper", None)
    if paper:
        details.append(str(paper))
    print_sides = getattr(item, "print_sides", None)
    if print_sides:
        details.append(str(print_sides))
    material = getattr(item, "material", None)
    if material:
        details.append(str(material))

    details.extend(_dynamic_spec_details(item))

    return ", ".join(dict.fromkeys(details))


def _rich_item_text(item: Any, base_font_name: str | None) -> CellRichText:
    product = str(item.product_name or "품목").strip()
    detail = _detail_text(item)
    rich = CellRichText()
    rich.append(TextBlock(InlineFont(rFont=base_font_name, b=True, sz=14), product))
    if detail:
        rich.append("\n")
        rich.append(TextBlock(InlineFont(rFont=base_font_name, b=False, sz=9), f"({detail})"))
    return rich


def _sheet_base_name(mail: Mail, now: datetime) -> str:
    person = _safe_name(mail.customer_name, "담당자")
    raw = f"{now:%m%d}_{person}"
    return re.sub(r"[\\/*?:\[\]]", "_", raw)[:31]


def _unique_sheet_name(workbook, base: str, ignored: str | None = None) -> str:
    names = {name for name in workbook.sheetnames if name != ignored}
    if base not in names:
        return base
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in names:
            return candidate
        index += 1


def _find_mail_sheet(workbook, mail_id: int):
    marker = f"{MAIL_MARKER_PREFIX}{mail_id}"
    for sheet in workbook.worksheets:
        max_column = min(sheet.max_column, 200)
        for row in range(1, min(sheet.max_row, 10) + 1):
            for column in range(1, max_column + 1):
                cell = sheet.cell(row, column)
                if not isinstance(cell, MergedCell) and cell.value == marker:
                    return sheet
    return None


def _write_mail_marker(sheet, mail_id: int) -> None:
    column = max(sheet.max_column + 1, 20)
    while column <= 16384:
        cell = sheet.cell(1, column)
        if not isinstance(cell, MergedCell):
            cell.value = f"{MAIL_MARKER_PREFIX}{mail_id}"
            sheet.column_dimensions[get_column_letter(column)].hidden = True
            return
        column += 1
    raise ValueError("견적 시트 내부 식별값을 기록할 안전한 셀을 찾지 못했습니다.")


def _populate_sheet(sheet, settings: Settings, mail: Mail, selected: list[QuotationDraftItem], total: int, complete: bool, now: datetime) -> None:
    customer_name = mail.customer_organization or mail.customer_name or "고객"
    _set_value(sheet, "B3", f"{customer_name} 귀하")
    _set_value(sheet, "D4", mail.customer_department or "담당자 귀하")
    _set_value(sheet, "D5", now)
    _set_value(sheet, "D6", mail.delivery_place or settings.default_delivery_place)
    _set_value(sheet, "D7", mail.payment_terms or settings.default_payment_terms)
    _set_value(sheet, "D8", settings.default_validity)
    _set_value(sheet, "L5", mail.customer_name or "")
    _set_value(sheet, "L6", mail.customer_phone or "")
    _set_value(sheet, "L7", mail.customer_email or mail.original_sender_email or "")
    _clear_item_area(sheet)
    for row, (draft_item, mail_item) in enumerate(zip(selected, mail.items), start=ITEM_START_ROW):
        if row > ITEM_END_ROW:
            break
        _merged_anchor(sheet, row, 2).value = row - ITEM_START_ROW + 1
        item_cell = _merged_anchor(sheet, row, 3)
        item_cell.value = _rich_item_text(mail_item, item_cell.font.name)
        _merged_anchor(sheet, row, 6).value = draft_item.quantity
        _merged_anchor(sheet, row, 7).value = draft_item.unit_price
        _merged_anchor(sheet, row, 9).value = draft_item.amount
        _merged_anchor(sheet, row, 12).value = draft_item.note

        # 제작 원가는 내부 관리용 숨김 열(T)에만 기록한다.
        _merged_anchor(sheet, row, 20).value = draft_item.cost_price

    sheet.column_dimensions["T"].hidden = True

    value = total if complete else ""
    _set_value(sheet, "G24", value)
    _set_value(sheet, "D10", value)
    _set_value(sheet, "I10", value)


def _backup_path(path: Path, now: datetime) -> Path:
    return path.with_name(f".{path.stem}.{now:%Y%m%d_%H%M%S}.{uuid.uuid4().hex}.backup.xlsx")


def _temporary_path(path: Path) -> Path:
    return path.with_name(f".{path.stem}.{uuid.uuid4().hex}.saving.xlsx")


def _requires_native_excel(path: Path) -> bool:
    """Detect package features that openpyxl cannot safely round-trip."""
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
    except (OSError, zipfile.BadZipFile):
        return False
    return any(
        name.startswith("xl/embeddings/")
        or "/vmlDrawing" in name
        or name.lower().endswith(".emf")
        or name.startswith("xl/ctrlProps/")
        or name.startswith("xl/activeX/")
        for name in names
    )


def _save_with_native_excel(
    temp_path: Path,
    settings: Settings,
    mail: Mail,
    selected: list[QuotationDraftItem],
    total: int,
    complete: bool,
    now: datetime,
) -> None:
    if sys.platform != "win32":
        raise RuntimeError("복합 개체가 포함된 견적 파일은 Windows용 Microsoft Excel이 필요합니다.")
    token = uuid.uuid4().hex
    script_path = temp_path.with_name(f".{temp_path.stem}.{token}.excel-save.ps1")
    payload_path = temp_path.with_name(f".{temp_path.stem}.{token}.excel-save.json")
    value: int | str = total if complete else ""
    payload = {
        "marker": f"{MAIL_MARKER_PREFIX}{mail.id}",
        "base_name": _sheet_base_name(mail, now),
        "cells": {
            "B3": f"{mail.customer_organization or mail.customer_name or '고객'} 귀하",
            "D4": mail.customer_department or "담당자 귀하",
            "D5": now.isoformat(),
            "D6": mail.delivery_place or settings.default_delivery_place,
            "D7": mail.payment_terms or settings.default_payment_terms,
            "D8": settings.default_validity,
            "L5": mail.customer_name or "",
            "L6": mail.customer_phone or "",
            "L7": mail.customer_email or mail.original_sender_email or "",
            "G24": value,
            "D10": value,
            "I10": value,
        },
        "items": [
            {
                "row": ITEM_START_ROW + index,
                "position": index + 1,
                "product": str(mail_item.product_name or "품목").strip(),
                "detail": _detail_text(mail_item),
                "quantity": draft_item.quantity,
                "unit_price": draft_item.unit_price,
                "amount": draft_item.amount,
                "note": draft_item.note,
                "cost_price": draft_item.cost_price,
            }
            for index, (draft_item, mail_item) in enumerate(zip(selected, mail.items))
            if ITEM_START_ROW + index <= ITEM_END_ROW
        ],
    }
    try:
        script_path.write_text(EXCEL_COM_SCRIPT, encoding="utf-8-sig")
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8-sig")
        result = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-NonInteractive",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-TargetPath",
                str(temp_path),
                "-TemplatePath",
                str(settings.quotation_template_path),
                "-TemplateSheet",
                settings.quotation_template_sheet,
                "-PayloadPath",
                str(payload_path),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
            check=False,
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "Microsoft Excel 자동화에 실패했습니다.").strip()
            raise RuntimeError(f"기존 견적 파일에 시트를 추가하지 못했습니다. Microsoft Excel을 확인해주세요.\n{detail}")
    except FileNotFoundError as error:
        raise RuntimeError("복합 개체가 포함된 견적 파일을 처리하려면 Microsoft Excel과 PowerShell이 필요합니다.") from error
    finally:
        script_path.unlink(missing_ok=True)
        payload_path.unlink(missing_ok=True)


def _atomic_replace(temp_path: Path, target_path: Path) -> None:
    try:
        os.replace(temp_path, target_path)
    except PermissionError as error:
        raise QuotationFileLockedError(
            "견적 엑셀 파일이 Excel 또는 다른 프로그램에서 열려 있습니다.\n"
            "해당 파일을 닫은 뒤 다시 견적서를 생성해주세요."
        ) from error
    except OSError as error:
        if getattr(error, "winerror", None) == 32:
            raise QuotationFileLockedError(
                "견적 엑셀 파일이 Excel 또는 다른 프로그램에서 열려 있습니다.\n"
                "해당 파일을 닫은 뒤 다시 견적서를 생성해주세요."
            ) from error
        raise


def create_quotation(
    session: Session,
    settings: Settings,
    mail: Mail,
    *,
    storage_mode: StorageMode = "separate",
    target_path: Path | None = None,
    now: datetime | None = None,
) -> QuotationDraft:
    now = now or datetime.now()
    _ensure_no_blocking_reviews(session, mail)
    if mail.analysis_payload.get("is_order_related") is False:
        raise ValueError("견적 업무와 관련된 메일만 견적서를 생성할 수 있습니다.")
    validation_errors = validate_quote_items(list(mail.items))
    if validation_errors:
        raise ValueError(
            "견적서 생성에 필요한 기본 정보를 확인해 주세요. "
            + " / ".join(validation_errors)
        )
    if not settings.quotation_template_path.exists():
        raise FileNotFoundError(f"견적서 템플릿을 찾을 수 없습니다: {settings.quotation_template_path}")
    if target_path is None:
        generated = {
            row["mode"]: Path(row["path"])
            for row in _new_file_candidates(settings, mail, now)
        }
        target_path = generated[storage_mode]
    target, existed = _validate_target(settings, mail, storage_mode, target_path, now)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp_path = _temporary_path(target)
    backup_path: Path | None = None
    replaced = False
    customer_name = mail.customer_organization or mail.customer_name or "고객"
    draft = session.scalar(
        select(QuotationDraft)
        .where(QuotationDraft.mail_id == mail.id)
        .order_by(QuotationDraft.id.desc())
        .options(selectinload(QuotationDraft.items))
    )
    if draft is None:
        draft = QuotationDraft(mail_id=mail.id, status=DraftStatus.DRAFT, file_path=str(target), customer_name=customer_name)
        session.add(draft)
    else:
        draft.items.clear()
        draft.status = DraftStatus.DRAFT
        draft.customer_name = customer_name
        draft.approved_at = draft.sent_at = draft.sent_to = draft.error_message = None
    draft.file_path = str(target)
    draft.email_subject = f"[열린문디자인] 요청하신 견적서를 보내드립니다 - {customer_name}"
    draft.email_body = (
        f"안녕하세요. {customer_name} 담당자님.\n\n요청하신 견적서를 첨부하여 보내드립니다.\n"
        "검토 후 문의사항이 있으시면 회신 부탁드립니다.\n\n감사합니다.\n열린문디자인"
    )

    target_workbook = template_workbook = None
    try:
        session.flush()
        total = 0
        complete = True
        selected: list[QuotationDraftItem] = []
        for position, item in enumerate(mail.items, start=1):
            unit_price, amount, source = _resolve_item_price(settings, mail, item)
            item.unit_price, item.amount = unit_price, amount
            if amount is None:
                complete = False
            else:
                total += amount
            record = QuotationDraftItem(
                draft_id=draft.id,
                position=position,
                product_name=item.product_name,
                specification=item.specification,
                spec_attributes=dict(item.spec_attributes or {}),
                quantity=item.quantity,
                unit=item.unit,
                cost_price=item.cost_price,
                unit_price=unit_price,
                amount=amount,
                note=item.schedule_note,
                price_source=source,
            )
            session.add(record)
            session.flush()
            selected.append(record)

        if existed:
            backup_path = _backup_path(target, now)
            shutil.copy2(target, backup_path)
        if existed and _requires_native_excel(target):
            shutil.copy2(target, temp_path)
            _save_with_native_excel(temp_path, settings, mail, selected, total, complete, now)
        else:
            template_workbook = load_workbook(settings.quotation_template_path, rich_text=True)
            template_sheet = (
                template_workbook[settings.quotation_template_sheet]
                if settings.quotation_template_sheet in template_workbook.sheetnames
                else template_workbook.worksheets[0]
            )
            if existed:
                target_workbook = load_workbook(target, rich_text=True)
            else:
                target_workbook = Workbook()
                target_workbook.remove(target_workbook.active)

            old_sheet = _find_mail_sheet(target_workbook, mail.id)
            old_index = target_workbook.worksheets.index(old_sheet) if old_sheet is not None else None
            old_title = old_sheet.title if old_sheet is not None else None
            base_name = _sheet_base_name(mail, now)
            sheet_name = old_title or _unique_sheet_name(target_workbook, base_name)
            if old_sheet is not None:
                target_workbook.remove(old_sheet)
            sheet = _copy_template_sheet(template_sheet, target_workbook, sheet_name, old_index)
            _populate_sheet(sheet, settings, mail, selected, total, complete, now)
            _write_mail_marker(sheet, mail.id)
            target_workbook.active = target_workbook.worksheets.index(sheet)
            try:
                target_workbook.calculation.fullCalcOnLoad = True
                target_workbook.calculation.forceFullCalc = True
                target_workbook.calculation.calcMode = "auto"
            except AttributeError:
                pass
            target_workbook.save(temp_path)
            target_workbook.close()
            target_workbook = None
        _atomic_replace(temp_path, target)
        replaced = True

        _export_customer_pdf(
            settings,
            draft,
            target,
            mail,
        )

        draft.total_amount = total if selected and complete else None
        mail.status = MailStatus.QUOTE_CREATED
        session.commit()
        session.refresh(draft)
        return draft
    except QuotationFileLockedError:
        session.rollback()
        temp_path.unlink(missing_ok=True)
        if replaced and existed and backup_path and backup_path.exists():
            shutil.copy2(backup_path, target)
        elif not existed:
            target.unlink(missing_ok=True)
        raise
    except PermissionError as error:
        session.rollback()
        temp_path.unlink(missing_ok=True)
        if replaced and existed and backup_path and backup_path.exists():
            shutil.copy2(backup_path, target)
        elif not existed:
            target.unlink(missing_ok=True)
        raise QuotationFileLockedError(
            "견적 엑셀 파일이 Excel 또는 다른 프로그램에서 열려 있습니다.\n"
            "해당 파일을 닫은 뒤 다시 견적서를 생성해주세요."
        ) from error
    except Exception:
        session.rollback()
        temp_path.unlink(missing_ok=True)
        if replaced and existed and backup_path and backup_path.exists():
            shutil.copy2(backup_path, target)
        elif not existed:
            target.unlink(missing_ok=True)
        raise
    finally:
        if target_workbook is not None:
            try:
                target_workbook.close()
            except Exception:
                pass
        if template_workbook is not None:
            try:
                template_workbook.close()
            except Exception:
                pass


def approve_draft(session: Session, draft: QuotationDraft) -> QuotationDraft:
    if draft.status != DraftStatus.DRAFT:
        raise ValueError("초안 상태의 견적서만 승인할 수 있습니다.")
    draft.status = DraftStatus.APPROVED
    draft.approved_at = datetime.now().astimezone().replace(tzinfo=None)
    draft.mail.status = MailStatus.APPROVED
    session.commit()
    session.refresh(draft)
    return draft
