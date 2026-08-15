from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from backend.app.config import Settings
from backend.app.database import Base
from backend.app.models import Mail, MailItem
from backend.app.services import quotation_service
from backend.app.services.quotation_service import (
    QuotationFileLockedError,
    create_quotation,
    get_storage_options,
)


NOW = datetime(2026, 8, 15, 10, 30)


@pytest.fixture()
def quote_environment(tmp_path: Path):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}")
    Base.metadata.create_all(engine)
    template = Path(__file__).resolve().parents[1] / "data" / "templates" / "quotation_template.xlsx"
    quote_root = tmp_path / "quotation_files"
    quote_root.mkdir()
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'test.db'}",
        quotation_template_path=template,
        quotation_files_path=quote_root,
        use_external_price_engine=False,
    )
    with Session(engine, expire_on_commit=False) as session:
        yield session, settings, quote_root, template


def _mail(session: Session, uid: str, person: str = "홍길동", product: str = "현수막") -> Mail:
    mail = Mail(
        uid=uid,
        customer_organization="아산시청",
        customer_department="홍보담당관",
        customer_name=person,
        customer_email="customer@example.com",
        status="READY_FOR_QUOTE",
        analysis_payload={"is_order_related": True},
    )
    mail.items = [
        MailItem(
            position=1,
            product_name=product,
            specification="가로 5000mm, 세로 600mm",
            width_mm=5000,
            height_mm=600,
            quantity=1,
            unit="장",
            paper="아트지",
            print_sides="단면",
            material="현수막",
            unit_price=33000,
            amount=33000,
            schedule_note="8월 20일 납품",
            confirmed=True,
        ),
        MailItem(
            position=2,
            product_name="배너",
            specification="600x1800mm",
            quantity=2,
            unit="개",
            unit_price=20000,
            amount=40000,
            confirmed=True,
        ),
    ]
    session.add(mail)
    session.commit()
    return mail


def _existing_file(template: Path, destination: Path) -> None:
    shutil.copy2(template, destination)
    workbook = load_workbook(destination)
    original = workbook["Main_Sheet"]
    original.title = "과거견적"
    original["C14"] = "과거 품목 - 절대 변경 금지"
    original["N2"] = "기존 통합문서 데이터"
    workbook.save(destination)
    workbook.close()


def test_adds_clean_template_sheet_and_preserves_existing_content(quote_environment):
    session, settings, root, template = quote_environment
    target = root / "26-아산시청 홍보담당관-홍길동.xlsx"
    _existing_file(template, target)
    mail = _mail(session, "mail-1")

    create_quotation(session, settings, mail, storage_mode="existing", target_path=target, now=NOW)

    workbook = load_workbook(target, rich_text=True)
    assert workbook["과거견적"]["C14"].value == "과거 품목 - 절대 변경 금지"
    assert workbook["과거견적"]["N2"].value == "기존 통합문서 데이터"
    sheet = workbook["0815_홍길동"]
    assert sheet["B14"].value == 1
    assert sheet["B15"].value == 2
    assert sheet["C16"].value is None
    assert sheet["C14"].value != "과거 품목 - 절대 변경 금지"
    assert len(sheet._images) == 4
    rich = sheet["C14"].value
    assert isinstance(rich, CellRichText)
    blocks = [part for part in rich if isinstance(part, TextBlock)]
    assert blocks[0].font.b is True and blocks[0].font.sz == 14
    assert blocks[-1].font.b is False and blocks[-1].font.sz == 9
    assert "(가로 5000mm, 세로 600mm, 5000*600mm, 1장, 아트지, 단면, 현수막)" in str(rich)
    workbook.close()
    assert list(root.glob(".*.backup.xlsx"))


def test_duplicate_names_and_same_mail_replaces_its_sheet(quote_environment):
    session, settings, root, template = quote_environment
    target = root / "26-아산시청 홍보담당관.xlsx"
    _existing_file(template, target)
    first = _mail(session, "mail-a")
    second = _mail(session, "mail-b")
    third = _mail(session, "mail-c")
    create_quotation(session, settings, first, storage_mode="existing", target_path=target, now=NOW)
    create_quotation(session, settings, second, storage_mode="existing", target_path=target, now=NOW)
    create_quotation(session, settings, third, storage_mode="existing", target_path=target, now=NOW)

    workbook = load_workbook(target)
    assert {"0815_홍길동", "0815_홍길동_2", "0815_홍길동_3"}.issubset(workbook.sheetnames)
    sheet_count = len(workbook.sheetnames)
    workbook.close()

    first.items[0].product_name = "수정된 현수막"
    session.commit()
    create_quotation(session, settings, first, storage_mode="existing", target_path=target, now=NOW)
    workbook = load_workbook(target, rich_text=True)
    assert len(workbook.sheetnames) == sheet_count
    assert str(workbook["0815_홍길동"]["C14"].value) == "수정된 현수막\n(가로 5000mm, 세로 600mm, 5000*600mm, 1장, 아트지, 단면, 현수막)"
    workbook.close()


def test_cross_workbook_style_copy_survives_save_and_reload(quote_environment):
    session, settings, root, template = quote_environment
    target = root / "26-아산시청.xlsx"
    _existing_file(template, target)
    mail = _mail(session, "mail-style")
    create_quotation(session, settings, mail, storage_mode="existing", target_path=target, now=NOW)
    source_book = load_workbook(template)
    result_book = load_workbook(target)
    source = source_book["Main_Sheet"]
    result = result_book["0815_홍길동"]
    assert result["A1"].font.name == source["A1"].font.name
    assert result["A1"].font.sz == source["A1"].font.sz
    assert result["B13"].fill.fill_type == source["B13"].fill.fill_type
    assert result["B13"].fill.fgColor.type == source["B13"].fill.fgColor.type
    assert result["G14"].border.left.style == source["G14"].border.left.style
    assert result.column_dimensions["C"].width == source.column_dimensions["C"].width
    assert result.row_dimensions[14].height == source.row_dimensions[14].height
    assert {str(value) for value in result.merged_cells.ranges} == {str(value) for value in source.merged_cells.ranges}
    result_book.save(root / "reopened-copy.xlsx")
    result_book.close()
    source_book.close()
    reopened = load_workbook(root / "reopened-copy.xlsx")
    assert "0815_홍길동" in reopened.sheetnames
    reopened.close()


def test_locked_file_message_and_failed_save_preserve_original(quote_environment, monkeypatch):
    session, settings, root, template = quote_environment
    target = root / "26-아산시청 잠김.xlsx"
    _existing_file(template, target)
    original = target.read_bytes()
    mail = _mail(session, "mail-locked")

    def locked_replace(_source, _target):
        raise PermissionError(13, "locked")

    monkeypatch.setattr(quotation_service.os, "replace", locked_replace)
    with pytest.raises(QuotationFileLockedError, match="Excel 또는 다른 프로그램에서 열려 있습니다"):
        create_quotation(session, settings, mail, storage_mode="existing", target_path=target, now=NOW)
    assert target.read_bytes() == original
    assert not list(root.glob(".*.saving.xlsx"))
    assert list(root.glob(".*.backup.xlsx"))
    assert session.query(quotation_service.QuotationDraft).count() == 0


def test_database_failure_after_replace_restores_original(quote_environment, monkeypatch):
    session, settings, root, template = quote_environment
    target = root / "26-아산시청 DB실패.xlsx"
    _existing_file(template, target)
    original = target.read_bytes()
    mail = _mail(session, "mail-db-failure")

    def failed_commit():
        raise RuntimeError("database commit failed")

    monkeypatch.setattr(session, "commit", failed_commit)
    with pytest.raises(RuntimeError, match="database commit failed"):
        create_quotation(session, settings, mail, storage_mode="existing", target_path=target, now=NOW)
    assert target.read_bytes() == original
    assert not list(root.glob(".*.saving.xlsx"))
    assert list(root.glob(".*.backup.xlsx"))


def test_new_file_failure_removes_incomplete_output(quote_environment, monkeypatch):
    session, settings, _root, _template = quote_environment
    mail = _mail(session, "mail-new-failure")
    options = get_storage_options(settings, mail, now=NOW)
    department = next(row for row in options["new_files"] if row["mode"] == "department")
    target = Path(department["path"])

    def failed_replace(_source, _target):
        raise OSError("replace failed")

    monkeypatch.setattr(quotation_service.os, "replace", failed_replace)
    with pytest.raises(OSError, match="replace failed"):
        create_quotation(session, settings, mail, storage_mode="department", target_path=target, now=NOW)
    assert not target.exists()
    assert not list(target.parent.glob(".*.saving.xlsx"))


def test_storage_candidates_and_new_department_file(quote_environment):
    session, settings, root, _template = quote_environment
    related = root / "25-아산시청 홍보담당관-기존담당자.xlsx"
    unrelated = root / "25-다른기관.xlsx"
    related.touch()
    unrelated.touch()
    mail = _mail(session, "mail-candidates")
    options = get_storage_options(settings, mail, now=NOW)
    assert [row["filename"] for row in options["existing_files"]] == [related.name]
    department = next(row for row in options["new_files"] if row["mode"] == "department")
    assert department["filename"] == "26-아산시청 홍보담당관.xlsx"
    related.unlink()
    unrelated.unlink()
    create_quotation(
        session,
        settings,
        mail,
        storage_mode="department",
        target_path=Path(department["path"]),
        now=NOW,
    )
    workbook = load_workbook(department["path"])
    assert workbook.sheetnames == ["0815_홍길동"]
    workbook.close()


def test_rejects_path_outside_quotation_folder(quote_environment):
    session, settings, root, _template = quote_environment
    mail = _mail(session, "mail-outside")
    with pytest.raises(ValueError, match="quotation_files 폴더 밖"):
        create_quotation(
            session,
            settings,
            mail,
            storage_mode="existing",
            target_path=root.parent / "outside.xlsx",
            now=NOW,
        )


def test_real_quotation_copy_add_save_and_reopen(quote_environment):
    session, settings, root, _template = quote_environment
    real_source = Path(__file__).resolve().parents[1] / "data" / "quotation_files" / "클린도시과-둔포신창 다문화특화거리.xlsx"
    if not real_source.exists():
        pytest.skip("실제 견적 검증용 파일이 없습니다.")
    target = root / real_source.name
    shutil.copy2(real_source, target)
    before = load_workbook(target, data_only=False)
    original_names = before.sheetnames[:]
    original_values = {sheet.title: (str(sheet["A1"].value), str(sheet["B3"].value)) for sheet in before.worksheets}
    original_images = {sheet.title: len(sheet._images) for sheet in before.worksheets}
    before.close()
    mail = _mail(session, "mail-real-copy")
    create_quotation(session, settings, mail, storage_mode="existing", target_path=target, now=NOW)
    reopened = load_workbook(target, rich_text=True)
    assert reopened.sheetnames[:-1] == original_names
    assert {name: (str(reopened[name]["A1"].value), str(reopened[name]["B3"].value)) for name in original_names} == original_values
    assert {name: len(reopened[name]._images) for name in original_names} == original_images
    assert reopened["0815_홍길동"]["B14"].value == 1
    reopened.close()


def test_complex_real_workbook_is_routed_to_native_excel():
    root = Path(__file__).resolve().parents[1] / "data" / "quotation_files"
    backups = list(root.glob(".26-충남사회경제네트워크*.backup.xlsx"))
    if not backups:
        pytest.skip("복합 개체 검증용 백업 파일이 없습니다.")
    assert quotation_service._requires_native_excel(backups[0]) is True
