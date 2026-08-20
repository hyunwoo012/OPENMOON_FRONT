from __future__ import annotations

from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.config import Settings, get_settings
from backend.app.database import Base, get_db
from backend.app.models import Mail
from backend.app.routers import mails, products, quotations


def _make_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Main_Sheet"

    sheet["B3"] = ""
    sheet["D4"] = ""
    sheet["D5"] = ""
    sheet["D6"] = ""
    sheet["D7"] = ""
    sheet["D8"] = ""
    sheet["D10"] = ""
    sheet["I10"] = ""
    sheet["G24"] = ""
    sheet["L5"] = ""
    sheet["L6"] = ""
    sheet["L7"] = ""

    workbook.save(path)
    workbook.close()


def _analysis_payload(items: list[dict]) -> dict:
    return {
        "customer_organization": "OPENMOON 1G 테스트기관",
        "customer_department": "테스트부서",
        "customer_name": "통합테스트",
        "customer_phone": "010-0000-0000",
        "customer_email": "integration@example.com",
        "delivery_place": "테스트 납품처",
        "payment_terms": "테스트 결제",
        "requested_date": "2026-08-31",
        "request_types": ["quotation"],
        "commitment_status": "unconfirmed",
        "summary": "1-G 통합 테스트",
        "reason": "동적 사양/원가/삭제/견적 스냅샷 테스트",
        "items": items,
    }


def test_product_catalog_full_roundtrip(tmp_path: Path):
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SessionLocal = sessionmaker(
        bind=engine,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
    )
    Base.metadata.create_all(bind=engine)

    template_path = tmp_path / "quotation_template.xlsx"
    quotation_dir = tmp_path / "quotation_files"
    quotation_dir.mkdir()
    _make_template(template_path)

    settings = Settings(
        database_url="sqlite://",
        quotation_template_path=template_path,
        quotation_template_sheet="Main_Sheet",
        quotation_files_path=quotation_dir,
        use_external_price_engine=False,
        allow_live_send=False,
    )

    with SessionLocal() as session:
        mail = Mail(
            account="1g-test",
            uid="1g-test-001",
            message_id="<openmoon-1g-test@example.com>",
            original_sender_name="통합테스트",
            original_sender_email="integration@example.com",
            original_subject="[1-G] 통합 테스트",
            original_body="동적 품목 사양 저장 테스트",
            customer_organization="OPENMOON 1G 테스트기관",
            customer_name="통합테스트",
            customer_email="integration@example.com",
            analysis_payload={"is_order_related": True},
        )
        session.add(mail)
        session.commit()
        mail_id = mail.id

    app = FastAPI()
    app.include_router(products.router)
    app.include_router(mails.router)
    app.include_router(quotations.router)

    def override_db():
        session = SessionLocal()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_settings] = lambda: settings

    client = TestClient(app)

    catalog_response = client.get("/api/products/catalog")
    assert catalog_response.status_code == 200
    catalog = catalog_response.json()
    assert sum(
        len(category["products"])
        for category in catalog["categories"]
    ) == 39

    name_card = {
        "product_name": "명함",
        "normalized_product": "명함",
        "specification": "90*50",
        "quantity": 200,
        "unit": "매",
        "paper": "휘라레",
        "print_sides": "단면4도",
        "spec_attributes": {
            "size": "90*50",
            "paper": "휘라레",
            "quantity": "200",
            "print_colors": "단면4도",
            "finishing": ["귀도리"],
            "delivery_method": "납품",
        },
        "cost_price": 18000,
        "unit_price": 150,
        "amount": 30000,
        "confirmed": True,
        "evidence": {},
    }

    banner = {
        "product_name": "현수막",
        "normalized_product": "현수막",
        "specification": "5000*900",
        "quantity": 2,
        "unit": "장",
        "spec_attributes": {
            "size": "5000*900",
            "finishing": "타공",
            "quantity": "2",
            "installation_delivery": "시공",
        },
        "cost_price": 35000,
        "unit_price": 60000,
        "amount": 120000,
        "confirmed": True,
        "evidence": {},
    }

    save_response = client.patch(
        f"/api/mails/{mail_id}/analysis",
        json=_analysis_payload([name_card, banner]),
    )
    assert save_response.status_code == 200, save_response.text
    saved = save_response.json()
    assert len(saved["items"]) == 2

    reload_response = client.get(f"/api/mails/{mail_id}")
    assert reload_response.status_code == 200
    reloaded = reload_response.json()
    assert len(reloaded["items"]) == 2

    by_name = {
        item["product_name"]: item
        for item in reloaded["items"]
    }

    assert by_name["명함"]["spec_attributes"]["finishing"] == ["귀도리"]
    assert by_name["명함"]["cost_price"] == 18000
    assert by_name["현수막"]["spec_attributes"]["finishing"] == "타공"
    assert by_name["현수막"]["spec_attributes"]["installation_delivery"] == "시공"
    assert by_name["현수막"]["cost_price"] == 35000

    delete_save = client.patch(
        f"/api/mails/{mail_id}/analysis",
        json=_analysis_payload([banner]),
    )
    assert delete_save.status_code == 200, delete_save.text

    after_delete = client.get(f"/api/mails/{mail_id}").json()
    assert len(after_delete["items"]) == 1
    assert after_delete["items"][0]["product_name"] == "현수막"
    assert after_delete["items"][0]["cost_price"] == 35000

    assert after_delete["status"] != "REVIEW_REQUIRED", after_delete.get("reviews")

    options_response = client.get(
        f"/api/quotations/storage-options/{mail_id}"
    )
    assert options_response.status_code == 200, options_response.text

    options = options_response.json()
    separate = next(
        row
        for row in options["new_files"]
        if row["mode"] == "separate"
    )

    quote_response = client.post(
        f"/api/quotations/from-mail/{mail_id}",
        json={
            "mode": "separate",
            "file_path": separate["path"],
        },
    )
    assert quote_response.status_code == 200, quote_response.text

    draft = quote_response.json()
    assert len(draft["items"]) == 1

    draft_item = draft["items"][0]
    assert draft_item["product_name"] == "현수막"
    assert draft_item["spec_attributes"]["finishing"] == "타공"
    assert draft_item["spec_attributes"]["installation_delivery"] == "시공"
    assert draft_item["cost_price"] == 35000
    assert draft["total_amount"] == 120000

    xlsx_path = Path(draft["file_path"])
    assert xlsx_path.exists()

    workbook = load_workbook(xlsx_path, rich_text=True)
    sheet = workbook.active

    assert sheet["T14"].value == 35000
    assert sheet.column_dimensions["T"].hidden is True
    assert sheet["G24"].value == 120000

    item_text = str(sheet["C14"].value)
    assert "마감처리: 타공" in item_text
    assert "시공여부: 시공" in item_text
    assert "35000" not in item_text

    workbook.close()
